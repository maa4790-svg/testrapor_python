# by Mahoni 2025-09-28
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Çoklu Sipariş Veri Çekme Scripti
"""

from main import PayzgateScraper
import time

def get_order_ids():
    """
    Kullanıcıdan sipariş ID'lerini alır
    
    Returns:
        list: Sipariş ID'leri listesi
    """
    print("📋 Sipariş ID'lerini Girin")
    print("=" * 40)
    print("Örnekler:")
    print("- Tek ID: 6172936")
    print("- Çoklu ID: 6172936,6172937,6172938")
    print("- Aralık: 6172936-6172940")
    print("- Karışık: 6172936,6172938-6172940,6172945")
    print()
    
    # Test için ID'ler koda yazıldı
    ids_input = "6172936,6172937,6172938"
    print(f"📋 Test ID'leri: {ids_input}")
    
    if not ids_input:
        print("❌ Hiç ID girilmedi!")
        return []
    
    order_ids = []
    
    try:
        # Virgülle ayrılmış kısımları işle
        parts = ids_input.split(',')
        
        for part in parts:
            part = part.strip()
            
            if '-' in part:
                # Aralık formatı: 6172936-6172940
                start, end = part.split('-', 1)
                start = int(start.strip())
                end = int(end.strip())
                
                for i in range(start, end + 1):
                    order_ids.append(str(i))
            else:
                # Tek ID
                order_ids.append(part)
        
        # Duplikatları kaldır ve sırala
        order_ids = sorted(list(set(order_ids)))
        
        print(f"✅ {len(order_ids)} adet sipariş ID'si alındı:")
        for i, order_id in enumerate(order_ids, 1):
            print(f"   {i}. {order_id}")
        
        return order_ids
        
    except ValueError as e:
        print(f"❌ Geçersiz format: {str(e)}")
        return []

def get_login_credentials():
    """
    Kullanıcıdan login bilgilerini alır
    
    Returns:
        tuple: (email, password)
    """
    print("\n🔐 Login Bilgileri")
    print("=" * 40)
    
    # Email ve şifre koda yazıldı
    email = "aras@pay.com"
    password = "0JHe0ET1FfoD"
    return email, password

def save_multiple_orders_to_excel(all_orders_data, filename=None):
    """
    Çoklu sipariş verilerini tek Excel dosyasına kaydeder
    
    Args:
        all_orders_data: Tüm sipariş verileri listesi
        filename: Dosya adı (opsiyonel)
    """
    if not all_orders_data:
        print("❌ Kaydedilecek veri yok!")
        return
    
    if filename is None:
        filename = f"multiple_orders_{int(time.time())}.xlsx"
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Yeni workbook oluştur
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sipariş Verileri"
        
        # Başlık satırı
        headers = [
            'Sipariş ID', 'URL', 'Çekilme Tarihi', 'Status', 'ID', 'Customer',
            'Order Subtotal', 'Order Total', 'Reference Code', 'IBAN'
        ]
        
        # Başlıkları yaz
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Verileri yaz
        row = 2
        for order_data in all_orders_data:
            if order_data and 'specific_data' in order_data:
                values = [
                    order_data['order_id'],
                    order_data['url'],
                    order_data['extracted_at'],
                    order_data['specific_data']['status'] or 'VERİ ÇEKİLEMEDİ',
                    order_data['specific_data']['id'] or 'VERİ ÇEKİLEMEDİ',
                    order_data['specific_data']['customer'] or 'VERİ ÇEKİLEMEDİ',
                    order_data['specific_data']['order_subtotal'] or 'VERİ ÇEKİLEMEDİ',
                    order_data['specific_data']['order_total'] or 'VERİ ÇEKİLEMEDİ',
                    order_data['specific_data']['reference_code'] or 'VERİ ÇEKİLEMEDİ',
                    order_data['specific_data']['iban'] or 'VERİ ÇEKİLEMEDİ'
                ]
                
                for col, value in enumerate(values, 1):
                    ws.cell(row=row, column=col, value=value)
                
                row += 1
        
        # Sütun genişliklerini ayarla
        column_widths = [15, 60, 20, 15, 15, 30, 25, 20, 40, 60]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
        
        # Dosyayı kaydet
        wb.save(filename)
        print(f"\n📁 Çoklu sipariş Excel dosyası kaydedildi: {filename}")
        
        # Özet bilgi
        total_orders = len(all_orders_data)
        successful_orders = len([d for d in all_orders_data if d and d.get('specific_data', {}).get('status') != 'VERİ ÇEKİLEMEDİ'])
        failed_orders = total_orders - successful_orders
        
        print(f"📊 ÖZET:")
        print(f"   Toplam sipariş: {total_orders}")
        print(f"   Başarılı: {successful_orders}")
        print(f"   Başarısız: {failed_orders}")
        
    except ImportError:
        print("❌ openpyxl kütüphanesi bulunamadı!")
    except Exception as e:
        print(f"❌ Excel kaydetme hatası: {str(e)}")

def main():
    """Ana fonksiyon"""
    print("🚀 Payzgate Çoklu Sipariş Veri Çekme Botu")
    print("=" * 50)
    
    # Sipariş ID'lerini al
    order_ids = get_order_ids()
    if not order_ids:
        return
    
    # Login bilgilerini al
    email, password = get_login_credentials()
    
    # Scraper oluştur
    scraper = PayzgateScraper(email=email, password=password)
    
    print(f"\n🚀 VERİ ÇEKME İŞLEMİ BAŞLIYOR")
    print("=" * 50)
    
    all_orders_data = []
    
    for i, order_id in enumerate(order_ids, 1):
        print(f"\n📋 [{i}/{len(order_ids)}] Sipariş ID: {order_id}")
        print("-" * 40)
        
        # Veri çek
        data = scraper.get_order_data(order_id)
        
        if data:
            all_orders_data.append(data)
            
            # Kısa özet göster
            specific_data = data['specific_data']
            status = specific_data['status'] or 'VERİ ÇEKİLEMEDİ'
            customer = specific_data['customer'] or 'VERİ ÇEKİLEMEDİ'
            
            if status != 'VERİ ÇEKİLEMEDİ':
                print(f"✅ Başarılı: {customer} - {status}")
            else:
                print(f"❌ Başarısız: Veri çekilemedi")
        else:
            print(f"❌ Başarısız: Veri çekilemedi")
            # Boş veri ekle
            all_orders_data.append(None)
        
        # Kısa bekleme (site yükünü azaltmak için)
        if i < len(order_ids):
            time.sleep(1)
    
    # Tüm verileri Excel'e kaydet
    print(f"\n📁 EXCEL DOSYASI OLUŞTURULUYOR")
    print("=" * 50)
    
    save_multiple_orders_to_excel(all_orders_data)
    
    print(f"\n🎉 İŞLEM TAMAMLANDI!")
    print("=" * 50)

if __name__ == "__main__":
    main()
