#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Payzgate Veri Çekme Botu
Bu script Payzgate sitesinden sipariş verilerini çeker.
"""

import requests
from bs4 import BeautifulSoup
import json
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
# import pandas as pd  # Pandas kurulum sorunu nedeniyle geçici olarak kapatıldı

class PayzgateScraper:
    def __init__(self, email=None, password=None):
        self.base_url = "https://payzgate.com"
        self.session = requests.Session()
        self.email = email
        self.password = password
        self.setup_session()
        
    def setup_session(self):
        """Session için gerekli header'ları ayarla"""
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            'Accept-Language': 'tr-TR,tr;q=0.9,en;q=0.8',
            'Accept-Encoding': 'gzip, deflate, br',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1',
        })
    
    def login(self):
        """
        Payzgate sitesine login yapar
        
        Returns:
            bool: Login başarılı ise True
        """
        if not self.email or not self.password:
            print("❌ Email ve şifre gerekli!")
            return False
        
        try:
            print("🔐 Login işlemi başlatılıyor...")
            
            # Login sayfasını al
            login_url = f"{self.base_url}/login"
            print(f"📄 Login sayfası alınıyor: {login_url}")
            
            response = self.session.get(login_url)
            print(f"📊 Sayfa yanıt kodu: {response.status_code}")
            
            if response.status_code != 200:
                print(f"❌ Login sayfası alınamadı: {response.status_code}")
                print(f"Yanıt içeriği: {response.text[:200]}...")
                return False
            
            # CSRF token'ı çıkar
            soup = BeautifulSoup(response.text, 'html.parser')
            csrf_token = soup.find('input', {'name': '__RequestVerificationToken'})
            
            if not csrf_token:
                print("❌ CSRF token bulunamadı!")
                print("Sayfa içeriği kontrol ediliyor...")
                
                # Sayfa içeriğini kontrol et
                if "login" in response.text.lower():
                    print("✅ Login sayfası yüklendi")
                else:
                    print("❌ Login sayfası yüklenmedi, farklı sayfa geldi")
                
                return False
            
            print("✅ CSRF token bulundu")
            
            # Login verilerini hazırla
            login_data = {
                'Email': self.email,
                'Password': self.password,
                'RememberMe': 'false',
                '__RequestVerificationToken': csrf_token.get('value')
            }
            
            print(f"📤 Login isteği gönderiliyor... Email: {self.email}")
            
            # Login isteği gönder
            login_response = self.session.post(login_url, data=login_data, allow_redirects=True)
            
            print(f"📊 Login yanıt kodu: {login_response.status_code}")
            print(f"🔗 Yönlendirilen URL: {login_response.url}")
            
            # Login başarı kontrolü
            if login_response.status_code == 200:
                # URL kontrolü
                if 'login' not in login_response.url.lower():
                    print("✅ Login başarılı! Ana sayfaya yönlendirildi")
                    return True
                else:
                    print("❌ Login başarısız! Hala login sayfasındayız")
                    
                    # Hata mesajlarını kontrol et
                    soup = BeautifulSoup(login_response.text, 'html.parser')
                    error_messages = soup.find_all(class_='field-validation-error')
                    if error_messages:
                        print("🚨 Hata mesajları:")
                        for error in error_messages:
                            print(f"   - {error.get_text(strip=True)}")
                    
                    return False
            else:
                print(f"❌ Login isteği başarısız: {login_response.status_code}")
                return False
                
        except Exception as e:
            print(f"💥 Login hatası: {str(e)}")
            import traceback
            traceback.print_exc()
            return False
    
    def selenium_login(self):
        """
        Selenium ile login yapar (alternatif yöntem)
        
        Returns:
            bool: Login başarılı ise True
        """
        if not self.email or not self.password:
            print("❌ Email ve şifre gerekli!")
            return False
        
        try:
            print("🤖 Selenium ile login başlatılıyor...")
            
            # Chrome seçenekleri
            chrome_options = Options()
            chrome_options.add_argument("--headless")  # Arka planda çalıştır
            chrome_options.add_argument("--no-sandbox")
            chrome_options.add_argument("--disable-dev-shm-usage")
            chrome_options.add_argument("--disable-gpu")
            chrome_options.add_argument("--window-size=1920,1080")
            
            driver = webdriver.Chrome(
                service=Service(ChromeDriverManager().install()),
                options=chrome_options
            )
            
            # Login sayfasına git
            login_url = f"{self.base_url}/login"
            print(f"📄 Login sayfasına gidiliyor: {login_url}")
            
            driver.get(login_url)
            
            # Sayfa yüklenmesini bekle
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.ID, "Email"))
            )
            
            print("✅ Login sayfası yüklendi")
            
            # Email gir
            email_field = driver.find_element(By.ID, "Email")
            email_field.clear()
            email_field.send_keys(self.email)
            print(f"📧 Email girildi: {self.email}")
            
            # Şifre gir
            password_field = driver.find_element(By.ID, "Password")
            password_field.clear()
            password_field.send_keys(self.password)
            print("🔑 Şifre girildi")
            
            # Login butonuna tıkla
            login_button = driver.find_element(By.CSS_SELECTOR, "input[type='submit']")
            login_button.click()
            print("🖱️ Login butonuna tıklandı")
            
            # Yönlendirme bekle
            time.sleep(3)
            
            # Başarı kontrolü
            current_url = driver.current_url
            print(f"🔗 Mevcut URL: {current_url}")
            
            if 'login' not in current_url.lower():
                print("✅ Selenium login başarılı!")
                
                # Session cookie'lerini al
                cookies = driver.get_cookies()
                for cookie in cookies:
                    self.session.cookies.set(cookie['name'], cookie['value'])
                
                driver.quit()
                return True
            else:
                print("❌ Selenium login başarısız!")
                
                # Hata mesajlarını kontrol et
                try:
                    error_elements = driver.find_elements(By.CLASS_NAME, "field-validation-error")
                    if error_elements:
                        print("🚨 Hata mesajları:")
                        for error in error_elements:
                            if error.text.strip():
                                print(f"   - {error.text.strip()}")
                except:
                    pass
                
                driver.quit()
                return False
                
        except Exception as e:
            print(f"💥 Selenium login hatası: {str(e)}")
            try:
                driver.quit()
            except:
                pass
            return False
    
    def selenium_get_order_data(self, order_id):
        """
        Selenium ile login yapıp sipariş verisi çeker
        
        Args:
            order_id (str): Sipariş ID'si
            
        Returns:
            dict: Çekilen veriler
        """
        url = f"{self.base_url}/Admin/Order/Edit/{order_id}"
        
        try:
            print(f"🤖 Selenium ile sipariş verisi çekiliyor: {url}")
            
            # Chrome seçenekleri
            chrome_options = Options()
            chrome_options.add_argument("--headless")
            chrome_options.add_argument("--no-sandbox")
            chrome_options.add_argument("--disable-dev-shm-usage")
            chrome_options.add_argument("--disable-gpu")
            chrome_options.add_argument("--window-size=1920,1080")
            
            driver = webdriver.Chrome(
                service=Service(ChromeDriverManager().install()),
                options=chrome_options
            )
            
            # Önce login yap
            if not self.selenium_login_with_driver(driver):
                print("❌ Selenium login başarısız!")
                driver.quit()
                return None
            
            print("✅ Login başarılı, sipariş sayfasına gidiliyor...")
            
            # Sipariş sayfasına git
            driver.get(url)
            
            # Sayfa yüklenmesini bekle
            WebDriverWait(driver, 15).until(
                EC.presence_of_element_located((By.TAG_NAME, "body"))
            )
            
            # Sayfa başlığını kontrol et
            page_title = driver.title
            print(f"📄 Sayfa başlığı: {page_title}")
            
            # Login sayfasına yönlendirilip yönlendirilmediğini kontrol et
            current_url = driver.current_url
            print(f"🔗 Mevcut URL: {current_url}")
            
            if 'login' in current_url.lower():
                print("❌ Login sayfasına yönlendirildi, veri çekilemedi!")
                driver.quit()
                return None
            
            # Sayfa kaynağını al
            page_source = driver.page_source
            driver.quit()
            
            # BeautifulSoup ile parse et
            soup = BeautifulSoup(page_source, 'html.parser')
            
            # Veri çekme işlemleri
            order_data = self.extract_order_data(soup, order_id)
            
            return order_data
            
        except Exception as e:
            print(f"💥 Selenium veri çekme hatası: {str(e)}")
            try:
                driver.quit()
            except:
                pass
            return None
    
    def selenium_login_with_driver(self, driver):
        """
        Mevcut driver ile login yapar
        
        Args:
            driver: Selenium WebDriver
            
        Returns:
            bool: Login başarılı ise True
        """
        try:
            print("🔐 Selenium ile login yapılıyor...")
            
            # Login sayfasına git
            login_url = f"{self.base_url}/login"
            print(f"📄 Login sayfasına gidiliyor: {login_url}")
            
            driver.get(login_url)
            
            # Sayfa yüklenmesini bekle
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.ID, "Email"))
            )
            
            print("✅ Login sayfası yüklendi")
            
            # Email gir
            email_field = driver.find_element(By.ID, "Email")
            email_field.clear()
            email_field.send_keys(self.email)
            print(f"📧 Email girildi: {self.email}")
            
            # Şifre gir
            password_field = driver.find_element(By.ID, "Password")
            password_field.clear()
            password_field.send_keys(self.password)
            print("🔑 Şifre girildi")
            
            # Login butonuna tıkla
            login_button = driver.find_element(By.CSS_SELECTOR, "input[type='submit']")
            login_button.click()
            print("🖱️ Login butonuna tıklandı")
            
            # Yönlendirme bekle
            time.sleep(3)
            
            # Başarı kontrolü
            current_url = driver.current_url
            print(f"🔗 Login sonrası URL: {current_url}")
            
            if 'login' not in current_url.lower():
                print("✅ Selenium login başarılı!")
                return True
            else:
                print("❌ Selenium login başarısız!")
                
                # Hata mesajlarını kontrol et
                try:
                    error_elements = driver.find_elements(By.CLASS_NAME, "field-validation-error")
                    if error_elements:
                        print("🚨 Hata mesajları:")
                        for error in error_elements:
                            if error.text.strip():
                                print(f"   - {error.text.strip()}")
                except:
                    pass
                
                return False
                
        except Exception as e:
            print(f"💥 Selenium login hatası: {str(e)}")
            return False
    
    def extract_specific_order_data(self, soup, order_id):
        """
        HTML'den spesifik sipariş verilerini çıkarır
        
        Args:
            soup: BeautifulSoup objesi
            order_id: Sipariş ID'si
            
        Returns:
            dict: Çıkarılan spesifik veriler
        """
        data = {
            'status': '',
            'id': '',
            'customer': '',
            'order_subtotal': '',
            'order_total': '',
            'reference_code': '',
            'iban': ''
        }
        
        try:
            # Status - OrderStatus label'ından sonraki form-text-row
            status_label = soup.find('label', {'for': 'OrderStatus'})
            if status_label:
                status_div = status_label.find_next('div', class_='form-text-row')
                if status_div:
                    data['status'] = status_div.get_text(strip=True)
            
            # ID - CustomOrderNumber label'ından sonraki form-text-row
            id_label = soup.find('label', {'for': 'CustomOrderNumber'})
            if id_label:
                id_div = id_label.find_next('div', class_='form-text-row')
                if id_div:
                    data['id'] = id_div.get_text(strip=True)
            
            # Customer - Customer label'ından sonraki link
            customer_label = soup.find('label', {'for': 'CustomerId'})
            if customer_label:
                customer_link = customer_label.find_next('a')
                if customer_link:
                    data['customer'] = customer_link.get_text(strip=True)
            
            # Order Subtotal - Farklı yöntemlerle dene
            subtotal_found = False
            
            # Yöntem 1: "Order subtotal" label'ından sonraki form-text-row
            subtotal_label = soup.find('label', string=lambda text: text and 'Order subtotal' in text)
            if subtotal_label:
                subtotal_div = subtotal_label.find_next('div', class_='form-text-row')
                if subtotal_div:
                    data['order_subtotal'] = subtotal_div.get_text(strip=True)
                    subtotal_found = True
            
            # Yöntem 2: "Order subtotal" text'i içeren div'i ara
            if not subtotal_found:
                subtotal_divs = soup.find_all('div', class_='form-text-row')
                for div in subtotal_divs:
                    parent_div = div.find_parent('div', class_='form-group')
                    if parent_div:
                        label = parent_div.find('label')
                        if label and 'Order subtotal' in label.get_text():
                            data['order_subtotal'] = div.get_text(strip=True)
                            subtotal_found = True
                            break
            
            # Yöntem 3: "subtotal" kelimesi içeren form-text-row'u ara
            if not subtotal_found:
                subtotal_divs = soup.find_all('div', class_='form-text-row')
                for div in subtotal_divs:
                    text = div.get_text(strip=True)
                    if '₺' in text and ('excl tax' in text or 'incl tax' in text):
                        # Bu muhtemelen subtotal
                        data['order_subtotal'] = text
                        subtotal_found = True
                        break
            
            # Yöntem 4: Order Total ile aynı değeri kullan (subtotal yoksa)
            if not subtotal_found:
                order_total = data.get('order_total', '')
                if order_total and '₺' in order_total:
                    data['order_subtotal'] = f"{order_total} (Total ile aynı)"
                    subtotal_found = True
            
            # Yöntem 5: Hiç bulunamadıysa "BİLGİ YOK" yaz
            if not subtotal_found:
                data['order_subtotal'] = "BİLGİ YOK"
            
            # Order Total - OrderTotal label'ından sonraki form-text-row
            total_label = soup.find('label', {'for': 'OrderTotal'})
            if total_label:
                total_div = total_label.find_next('div', class_='form-text-row')
                if total_div:
                    data['order_total'] = total_div.get_text(strip=True)
            
            # Reference Code - BankReferenceCode label'ından sonraki form-text-row
            ref_label = soup.find('label', {'for': 'BankReferenceCode'})
            if ref_label:
                ref_div = ref_label.find_next('div', class_='form-text-row')
                if ref_div:
                    data['reference_code'] = ref_div.get_text(strip=True)
            
            # IBAN - IBAN label'ından sonraki form-text-row
            iban_label = soup.find('label', {'for': 'IBAN'})
            if iban_label:
                iban_div = iban_label.find_next('div', class_='form-text-row')
                if iban_div:
                    data['iban'] = iban_div.get_text(strip=True)
            
        except Exception as e:
            print(f"Spesifik veri çıkarma hatası: {str(e)}")
        
        # Debug mesajları kaldırıldı - production için
        
        return data
    
    def get_order_data(self, order_id):
        """
        Belirtilen sipariş ID'si için veri çeker
        
        Args:
            order_id (str): Sipariş ID'si
            
        Returns:
            dict: Çekilen veriler
        """
        url = f"{self.base_url}/Admin/Order/Edit/{order_id}"
        
        try:
            print(f"Sipariş verisi çekiliyor: {url}")
            
            # Önce login kontrolü yap
            if self.email and self.password:
                print("Login kontrolü yapılıyor...")
                if not self.login():
                    print("Requests ile login başarısız, Selenium ile denenecek...")
                    if not self.selenium_login():
                        print("Selenium ile de login başarısız, veri çekilemedi!")
                        return None
            
            # Login başarılı ise cookie'leri kullan, değilse Selenium ile login yap
            if self.email and self.password:
                # Selenium ile login ve veri çekme
                order_data = self.selenium_get_order_data(order_id)
                return order_data
            else:
                # Login olmadan veri çekme
                chrome_options = Options()
                chrome_options.add_argument("--headless")  # Arka planda çalıştır
                chrome_options.add_argument("--no-sandbox")
                chrome_options.add_argument("--disable-dev-shm-usage")
                chrome_options.add_argument("--disable-gpu")
                chrome_options.add_argument("--window-size=1920,1080")
                
                driver = webdriver.Chrome(
                    service=Service(ChromeDriverManager().install()),
                    options=chrome_options
                )
                
                driver.get(url)
                
                # Sayfanın yüklenmesini bekle
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.TAG_NAME, "body"))
                )
                
                # Sayfa kaynağını al
                page_source = driver.page_source
                driver.quit()
                
                # BeautifulSoup ile parse et
                soup = BeautifulSoup(page_source, 'html.parser')
                
                # Veri çekme işlemleri
                order_data = self.extract_order_data(soup, order_id)
                
                return order_data
            
        except Exception as e:
            print(f"Hata oluştu: {str(e)}")
            return None
    
    def extract_order_data(self, soup, order_id):
        """
        HTML'den sipariş verilerini çıkarır
        
        Args:
            soup: BeautifulSoup objesi
            order_id: Sipariş ID'si
            
        Returns:
            dict: Çıkarılan veriler
        """
        # Spesifik verileri çıkar
        order_data = self.extract_specific_order_data(soup, order_id)
        
        data = {
            'order_id': order_id,
            'url': f"{self.base_url}/Admin/Order/EditSimple/{order_id}",
            'extracted_at': time.strftime('%Y-%m-%d %H:%M:%S'),
            'specific_data': order_data,
            'raw_html': str(soup),
            'forms': [],
            'tables': [],
            'inputs': [],
            'text_content': soup.get_text()[:1000]  # İlk 1000 karakter
        }
        
        # Formları bul
        forms = soup.find_all('form')
        for form in forms:
            form_data = {
                'action': form.get('action', ''),
                'method': form.get('method', ''),
                'inputs': []
            }
            
            inputs = form.find_all(['input', 'select', 'textarea'])
            for inp in inputs:
                input_data = {
                    'name': inp.get('name', ''),
                    'type': inp.get('type', ''),
                    'value': inp.get('value', ''),
                    'id': inp.get('id', '')
                }
                form_data['inputs'].append(input_data)
            
            data['forms'].append(form_data)
        
        # Tabloları bul
        tables = soup.find_all('table')
        for table in tables:
            table_data = []
            rows = table.find_all('tr')
            for row in rows:
                cells = row.find_all(['td', 'th'])
                row_data = [cell.get_text(strip=True) for cell in cells]
                table_data.append(row_data)
            data['tables'].append(table_data)
        
        # Tüm inputları bul
        all_inputs = soup.find_all(['input', 'select', 'textarea'])
        for inp in all_inputs:
            input_data = {
                'name': inp.get('name', ''),
                'type': inp.get('type', ''),
                'value': inp.get('value', ''),
                'id': inp.get('id', ''),
                'class': inp.get('class', [])
            }
            data['inputs'].append(input_data)
        
        return data
    
    def save_data(self, data, filename=None):
        """
        Veriyi dosyaya kaydeder
        
        Args:
            data: Kaydedilecek veri
            filename: Dosya adı (opsiyonel)
        """
        if filename is None:
            filename = f"order_{data['order_id']}_{int(time.time())}.json"
        
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        
        print(f"Veri kaydedildi: {filename}")
    
    def save_to_excel(self, data, filename=None):
        """
        Veriyi Excel dosyasına kaydeder (tek formda)
        
        Args:
            data: Kaydedilecek veri
            filename: Dosya adı (opsiyonel)
        """
        if filename is None:
            filename = f"order_{data['order_id']}_{int(time.time())}.xlsx"
        
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Yeni workbook oluştur
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sipariş Verileri"
            
            # Başlık satırı
            headers = [
                'Sipariş ID', 'URL', 'Çekilme Tarihi', 'Status', 'ID', 'Customer',
                'Order Subtotal', 'Order Total', 'Reference Code', 'IBAN'
            ]
            
            # Veri satırı
            values = [
                data['order_id'],
                data['url'],
                data['extracted_at'],
                data['specific_data']['status'] or 'VERİ ÇEKİLEMEDİ',
                data['specific_data']['id'] or 'VERİ ÇEKİLEMEDİ',
                data['specific_data']['customer'] or 'VERİ ÇEKİLEMEDİ',
                data['specific_data']['order_subtotal'] or 'VERİ ÇEKİLEMEDİ',
                data['specific_data']['order_total'] or 'VERİ ÇEKİLEMEDİ',
                data['specific_data']['reference_code'] or 'VERİ ÇEKİLEMEDİ',
                data['specific_data']['iban'] or 'VERİ ÇEKİLEMEDİ'
            ]
            
            # Başlıkları yaz
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Verileri yaz
            for col, value in enumerate(values, 1):
                ws.cell(row=2, column=col, value=value)
            
            # Sütun genişliklerini ayarla
            column_widths = [15, 60, 20, 15, 15, 30, 25, 20, 40, 60]
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
            
            # Dosyayı kaydet
            wb.save(filename)
            print(f"Excel dosyası kaydedildi: {filename}")
            
        except ImportError:
            print("openpyxl kütüphanesi bulunamadı. CSV formatında kaydediliyor...")
            self.save_to_csv_fallback(data, filename)
        except Exception as e:
            print(f"Excel kaydetme hatası: {str(e)}")
            self.save_to_csv_fallback(data, filename)
    
    def save_to_csv_fallback(self, data, filename):
        """
        Excel kaydetme başarısız olursa CSV olarak kaydet
        """
        import csv
        
        csv_filename = filename.replace('.xlsx', '.csv')
        
        # CSV için veri hazırla
        headers = [
            'Sipariş ID', 'URL', 'Çekilme Tarihi', 'Status', 'ID', 'Customer',
            'Order Subtotal', 'Order Total', 'Reference Code', 'IBAN'
        ]
        
        values = [
            data['order_id'],
            data['url'],
            data['extracted_at'],
            data['specific_data']['status'] or 'VERİ ÇEKİLEMEDİ',
            data['specific_data']['id'] or 'VERİ ÇEKİLEMEDİ',
            data['specific_data']['customer'] or 'VERİ ÇEKİLEMEDİ',
            data['specific_data']['order_subtotal'] or 'VERİ ÇEKİLEMEDİ',
            data['specific_data']['order_total'] or 'VERİ ÇEKİLEMEDİ',
            data['specific_data']['reference_code'] or 'VERİ ÇEKİLEMEDİ',
            data['specific_data']['iban'] or 'VERİ ÇEKİLEMEDİ'
        ]
        
        with open(csv_filename, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(headers)
            writer.writerow(values)
        
        print(f"CSV dosyası kaydedildi: {csv_filename}")

def main():
    """Ana fonksiyon"""
    print("Payzgate Veri Çekme Botu Başlatılıyor...")
    
    # Email ve şifre koda yazıldı
    email = "aras@pay.com"
    password = "0JHe0ET1FfoD"
    
    scraper = PayzgateScraper(email=email, password=password)
    
    # Sipariş ID'lerini al
    print("\n📋 Sipariş ID'lerini Girin")
    print("Örnekler:")
    print("- Tek ID: 6172936")
    print("- Çoklu ID: 6172936,6172937,6172938")
    print("- Aralık: 6172936-6172940")
    
    try:
        ids_input = input("Sipariş ID'lerini girin: ").strip()
        
        if not ids_input:
            print("❌ Hiç ID girilmedi!")
            return
        
        # ID'leri parse et
        order_ids = []
        parts = ids_input.split(',')
        
        for part in parts:
            part = part.strip()
            
            if '-' in part:
                # Aralık formatı
                start, end = part.split('-', 1)
                start = int(start.strip())
                end = int(end.strip())
                
                for i in range(start, end + 1):
                    order_ids.append(str(i))
            else:
                # Tek ID
                order_ids.append(part)
        
        # Duplikatları kaldır
        order_ids = list(set(order_ids))
        
        print(f"✅ {len(order_ids)} adet sipariş ID'si alındı")
        
        all_orders_data = []
        
        for i, order_id in enumerate(order_ids, 1):
            print(f"\n📋 [{i}/{len(order_ids)}] Sipariş ID: {order_id}")
            
            # Veri çek
            data = scraper.get_order_data(order_id)
            
            if data:
                all_orders_data.append(data)
                print("✅ Veri başarıyla çekildi!")
                
                # Spesifik verileri göster
                specific_data = data['specific_data']
                status = specific_data['status'] or 'VERİ ÇEKİLEMEDİ'
                customer = specific_data['customer'] or 'VERİ ÇEKİLEMEDİ'
                
                print(f"   Status: {status}")
                print(f"   Customer: {customer}")
            else:
                all_orders_data.append(None)
                print("❌ Veri çekilemedi!")
        
        # Tüm verileri tek Excel'e kaydet
        if all_orders_data:
            print(f"\n📁 EXCEL DOSYASI OLUŞTURULUYOR")
            save_multiple_orders_to_excel(all_orders_data)
        
    except EOFError:
        print("Input hatası, tek ID ile test ediliyor...")
        order_id = "6172936"
        data = scraper.get_order_data(order_id)
        
        if data:
            scraper.save_to_excel(data)
            print("Excel dosyası oluşturuldu!")
        else:
            print("Veri çekilemedi!")
    except Exception as e:
        print(f"Hata: {str(e)}")

def save_multiple_orders_to_excel(all_orders_data, filename=None):
    """
    Çoklu sipariş verilerini tek Excel dosyasına kaydeder
    """
    if not all_orders_data:
        print("❌ Kaydedilecek veri yok!")
        return
    
    if filename is None:
        filename = f"multiple_orders_{int(time.time())}.xlsx"
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Yeni workbook oluştur
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sipariş Verileri"
        
        # Başlık satırı
        headers = [
            'Sipariş ID', 'URL', 'Çekilme Tarihi', 'Status', 'ID', 'Customer',
            'Order Subtotal', 'Order Total', 'Reference Code', 'IBAN'
        ]
        
        # Başlıkları yaz
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Verileri yaz
        row = 2
        for order_data in all_orders_data:
            if order_data and 'specific_data' in order_data:
                values = [
                    order_data['order_id'],
                    order_data['url'],
                    order_data['extracted_at'],
                    order_data['specific_data']['status'] or 'VERİ ÇEKİLEMEDİ',
                    order_data['specific_data']['id'] or 'VERİ ÇEKİLEMEDİ',
                    order_data['specific_data']['customer'] or 'VERİ ÇEKİLEMEDİ',
                    order_data['specific_data']['order_subtotal'] or 'VERİ ÇEKİLEMEDİ',
                    order_data['specific_data']['order_total'] or 'VERİ ÇEKİLEMEDİ',
                    order_data['specific_data']['reference_code'] or 'VERİ ÇEKİLEMEDİ',
                    order_data['specific_data']['iban'] or 'VERİ ÇEKİLEMEDİ'
                ]
                
                for col, value in enumerate(values, 1):
                    ws.cell(row=row, column=col, value=value)
                
                row += 1
        
        # Sütun genişliklerini ayarla
        column_widths = [15, 60, 20, 15, 15, 30, 25, 20, 40, 60]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
        
        # Dosyayı kaydet
        wb.save(filename)
        print(f"📁 Çoklu sipariş Excel dosyası kaydedildi: {filename}")
        
        # Özet bilgi
        total_orders = len(all_orders_data)
        successful_orders = len([d for d in all_orders_data if d and d.get('specific_data', {}).get('status') != 'VERİ ÇEKİLEMEDİ'])
        failed_orders = total_orders - successful_orders
        
        print(f"📊 ÖZET:")
        print(f"   Toplam sipariş: {total_orders}")
        print(f"   Başarılı: {successful_orders}")
        print(f"   Başarısız: {failed_orders}")
        
    except ImportError:
        print("❌ openpyxl kütüphanesi bulunamadı!")
    except Exception as e:
        print(f"❌ Excel kaydetme hatası: {str(e)}")

if __name__ == "__main__":
    main()
